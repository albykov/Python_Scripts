#author: albykov@gmail.com
#non standard includes: openpyxl
#usage: scans excel spreadsheet and fills the gaps using previous rows

#load library
from openpyxl import load_workbook
import myhelpers
import fileLocations

def getShortNameNoSpaces(tn):
    if len(tn.split('+')) > 1:
        return tn.split('+')[0].replace(' ', '') + tn.split('+')[1].zfill(3)
    else:
        return tn.replace(' ', '')

#load workbook to memory
wbPath = fileLocations.lethbridge_TitleExcel
wb = load_workbook(wbPath)
#print wb.get_sheet_names()

#get active worksheet
ws = wb.active

#go through all rows
isFirstRow = True
for row in ws.iter_rows():
    if not isFirstRow:
        takePrevVal = 0
        #go through all cell of current row
        for cell in row:
            #if this is a first column
            if cell.column == 'A':
                #if this column is non None
                if cell.value is not None:
                    #save it to variable curr_row
                    curr_row = row
                    print cell.row
                    #cell.value = getShortNameNoSpaces(cell.value)
                    #if this cell is None
                elif cell.value is None:
                    #switch to fix it
                    takePrevVal = 1
                    #print '------------None:' + str(cell.row)
                    #if fix switcher is on
            if takePrevVal:
                #if current cell value is None
                if cell.value is None:
                    #print 'Col: ' + str(cell.col_idx)
                    #print 'Val+: ' + str(curr_row[cell.col_idx-1].value)

                    #if saved prev row cell value is non None
                    if curr_row[cell.col_idx-1].value is not None:
                        cell.value = curr_row[cell.col_idx-1].value

                    #if the column is TitleNum
                    if cell.column == 'A':
                        tn = curr_row[cell.col_idx-1].value
                    #if the column is LINC
                    elif cell.column == 'B':
                        ln = curr_row[cell.col_idx-1].value
                else:
                    #print 'Col: ' + str(cell.col_idx)
                    #print 'Val: ' + str(cell.value)
                    pass
            else:
                #if the column is TitleNum
                if cell.column == 'A':
                    tn = cell.value
                #if the column is LINC
                elif cell.column == 'B':
                    ln = cell.value

            #if it is Title Number or LINC
            if cell.column == 'S':
                cell.value = getShortNameNoSpaces(tn)
            elif cell.column == 'T':
                cell.value = getShortNameNoSpaces(ln)

            #if it is a covenant columt
            elif cell.column == 'R':
                er = 0
                ea = 0
                rc = 0
                row = 0
                rowd = 0
                rowde = 0
                ca = 0
                cad = 0
                cade = 0
                z = 0
                io = 0
                hi = 0
                #like type of seatch, case insensitive and the top if will overwrithe the bottom
                if 1==2:
                    pass

                elif 'caveat' in str(cell.value).lower():
                    #print 'ca+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    ca = 1
                elif 'discharge' in str(cell.value).lower():
                    if 1==2:
                        pass
                    elif 'right of' in str(cell.value).lower():
                        #print 'disROW+++++++++++++++++++++++++++++++++++'+str(cell.value)
                        if 'except' in str(cell.value).lower():
                            #print 'disROWex+++++++++++++++++++++++++++++++++++'+str(cell.value)
                            rowde = 1
                        else:
                            #print 'disROW+++++++++++++++++++++++++++++++++++'+str(cell.value)
                            rowd = 1
                    elif 'caveat' in str(cell.value).lower():
                        if 'except' in str(cell.value).lower():
                            #print 'disCAex+++++++++++++++++++++++++++++++++++'+str(cell.value)
                            cade = 1
                        else:
                            #print 'disCA+++++++++++++++++++++++++++++++++++'+str(cell.value)
                            cad = 1
                elif 'environm' in str(cell.value).lower():
                    #print 'er+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    er = 1
                elif 'easement' in str(cell.value).lower():
                    #print 'ea++++++++++++++++++++++++++++++++++++'+str(cell.value)
                    ea = 1
                elif 'restrict' in str(cell.value).lower():
                    #print 'rc+++++++++++++++++++++++++++++++++++++'+str(cell.value)
                    rc = 1
                elif 'right of' in str(cell.value).lower():
                    #print 'row+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    row = 1
                elif 'zon' in str(cell.value).lower():
                    #print 'zone+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    z = 1
                elif 'irrigation order' in str(cell.value).lower():
                    #print 'io+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    io = 1
                elif 'histor' in str(cell.value).lower():
                    #print 'hi+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    hi = 1
                else:
                    print '+++++++++++++++++++++++++++++++++++'+str(cell.value)
                    pass

            # if it is ER column
            elif cell.column == 'U':
                cell.value = er
            # if it is EA column
            elif cell.column == 'V':
                cell.value = ea
            # if it is RC column
            elif cell.column == 'W':
                cell.value = rc
            # if it is row column
            elif cell.column == 'X':
                cell.value = row
            # if it is z column
            elif cell.column == 'Y':
                cell.value = z
            # if it is io column
            elif cell.column == 'Z':
                cell.value = io
            elif cell.column == 'AA':
                cell.value = hi

            elif cell.column == 'AB':
                cell.value = rowd
            elif cell.column == 'AC':
                cell.value = rowde
            elif cell.column == 'AD':
                cell.value = ca
            elif cell.column == 'AE':
                cell.value = cad
            elif cell.column == 'AF':
                cell.value = cade

    else:
        isFirstRow = False
#save workbook
print myhelpers.getNewFilePathWithDateNoSpaces(wbPath)
wb.save(myhelpers.getNewFilePathWithDateNoSpaces(wbPath))